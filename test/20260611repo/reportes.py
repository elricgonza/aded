"""
reportes.py  — Blueprint de Reportes interactivos
Todas las rutas devuelven:
  • HTML completo  (formato=html, para carga AJAX en el panel)
  • Excel          (formato=excel)
  • PDF            (formato=pdf)
La página principal (index) es el shell con el menú lateral.
"""
import io
from datetime import datetime

from flask import (Blueprint, Response, abort, render_template,
                   request, send_file, url_for)
from flask_login import current_user, login_required
from sqlalchemy import func

from app import db
from app.models import (Alumno, Asignado, Curso, Inscrito,
                        Materia, Nota, Pago, Profesor)

reportes_bp = Blueprint('reportes', __name__, url_prefix='/reportes')

# ── Paleta ────────────────────────────────────────────────────────────────────
_PRIM   = (124, 31,  62)
_PRIM2  = (92,  21,  48)
_XL_HDR = 'FF7C1F3E'
_XL_ALT = 'FFFAE4EB'


# ═══════════════════════════════════════════════════════════════════════════════
# Helpers de datos
# ═══════════════════════════════════════════════════════════════════════════════

def _cursos():
    return Curso.query.order_by(Curso.gestion.desc(), Curso.paralelo).all()

def _alumnos_activos():
    return Alumno.query.filter_by(activo=True).order_by(Alumno.paterno, Alumno.nombre).all()

def _nc(curso):
    g = curso.grado.grado if curso and curso.grado else ''
    return f"{g} – {curso.paralelo} ({curso.gestion})" if curso else '—'

def _inscritos_curso(cur_id):
    return (Inscrito.query
            .join(Alumno, Inscrito.alu_id == Alumno.id)
            .filter(Inscrito.cur_id == cur_id,
                    db.or_(Inscrito.inscrito == True, Inscrito.reserva == True)))


# ═══════════════════════════════════════════════════════════════════════════════
# Helpers Excel
# ═══════════════════════════════════════════════════════════════════════════════

def _xl_styles():
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    thin   = Side(style='thin', color='FFB0B0B0')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    return dict(
        hf = PatternFill('solid', fgColor=_XL_HDR),
        af = PatternFill('solid', fgColor=_XL_ALT),
        hfont  = Font(bold=True, color='FFFFFFFF', size=10),
        bfont  = Font(size=9),
        boldf  = Font(bold=True, size=9),
        center = Alignment(horizontal='center', vertical='center', wrap_text=True),
        left   = Alignment(horizontal='left',   vertical='center', wrap_text=True),
        border = border,
    )

def _xl_hdr(ws, cols, s):
    ws.append(cols)
    for c in ws[1]:
        c.fill = s['hf']; c.font = s['hfont']
        c.border = s['border']; c.alignment = s['center']

def _xl_rows(ws, s, start=2):
    for i, row in enumerate(ws.iter_rows(min_row=start)):
        for c in row:
            if i % 2 == 0: c.fill = s['af']
            c.font = s['bfont']; c.border = s['border']; c.alignment = s['left']

def _xl_width(ws, lo=10, hi=40):
    for col in ws.columns:
        w = max((len(str(c.value or '')) for c in col), default=lo)
        ws.column_dimensions[col[0].column_letter].width = min(max(w+2, lo), hi)

def _xl_send(wb, name):
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, download_name=name, as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ═══════════════════════════════════════════════════════════════════════════════
# Helpers PDF
# ═══════════════════════════════════════════════════════════════════════════════

def _pdf_setup(landscape=False):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.pagesizes import landscape as RL_L
    from reportlab.platypus import SimpleDocTemplate
    buf = io.BytesIO()
    ps  = RL_L(A4) if landscape else A4
    doc = SimpleDocTemplate(buf, pagesize=ps,
                             leftMargin=36, rightMargin=36,
                             topMargin=36, bottomMargin=36)
    return doc, buf

def _pdf_title(title, subtitle=''):
    from reportlab.lib import colors
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, Spacer, Table, TableStyle
    r,g,b = _PRIM
    pc = colors.Color(r/255, g/255, b/255)
    ts = ParagraphStyle('T', fontSize=14, fontName='Helvetica-Bold',
                         textColor=colors.white, alignment=1)
    ss = ParagraphStyle('S', fontSize=9,  fontName='Helvetica',
                         textColor=colors.white, alignment=1)
    ds = ParagraphStyle('D', fontSize=7,  fontName='Helvetica',
                         textColor=colors.white, alignment=2)
    rows = [[Paragraph(title, ts)]]
    if subtitle: rows.append([Paragraph(subtitle, ss)])
    rows.append([Paragraph(f'Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")}', ds)])
    t = Table(rows, colWidths=['100%'])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0),(-1,-1), pc),
        ('TOPPADDING', (0,0),(-1,-1), 8),
        ('BOTTOMPADDING', (0,0),(-1,-1), 8),
        ('LEFTPADDING', (0,0),(-1,-1), 14),
        ('RIGHTPADDING', (0,0),(-1,-1), 14),
    ]))
    return [t, Spacer(1, 0.35*cm)]

def _pdf_tbl(data, widths, zebra=True):
    from reportlab.lib import colors
    from reportlab.platypus import Table, TableStyle
    r,g,b = _PRIM;  pc = colors.Color(r/255, g/255, b/255)
    r2,g2,b2 = (250,228,235); zc = colors.Color(r2/255,g2/255,b2/255)
    t = Table(data, colWidths=widths, repeatRows=1)
    st = [
        ('BACKGROUND',(0,0),(-1,0),pc),
        ('TEXTCOLOR',(0,0),(-1,0),colors.white),
        ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
        ('FONTSIZE',(0,0),(-1,0),8),
        ('ALIGN',(0,0),(-1,0),'CENTER'),
        ('FONTNAME',(0,1),(-1,-1),'Helvetica'),
        ('FONTSIZE',(0,1),(-1,-1),8),
        ('GRID',(0,0),(-1,-1),0.4,colors.Color(.75,.75,.75)),
        ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
        ('TOPPADDING',(0,0),(-1,-1),4),
        ('BOTTOMPADDING',(0,0),(-1,-1),4),
        ('LEFTPADDING',(0,0),(-1,-1),5),
        ('RIGHTPADDING',(0,0),(-1,-1),5),
    ]
    if zebra:
        for i in range(1, len(data)):
            if i%2==0: st.append(('BACKGROUND',(0,i),(-1,i),zc))
    t.setStyle(TableStyle(st))
    return t

def _pdf_send(doc, buf, story, name):
    doc.build(story); buf.seek(0)
    return send_file(buf, download_name=name, as_attachment=True,
                     mimetype='application/pdf')


# ═══════════════════════════════════════════════════════════════════════════════
# Página principal — shell interactivo
# ═══════════════════════════════════════════════════════════════════════════════

@reportes_bp.route('/')
@login_required
def index():
    return render_template('reportes/index.html',
                           cursos=_cursos(),
                           alumnos=_alumnos_activos())


# ═══════════════════════════════════════════════════════════════════════════════
# R1 – Alumnos por Curso
# ═══════════════════════════════════════════════════════════════════════════════

@reportes_bp.route('/r1')
@login_required
def r1():
    cur_id  = request.args.get('cur_id', type=int)
    orden   = request.args.get('orden', 'nombre')
    formato = request.args.get('formato', 'html')
    curso   = Curso.query.get(cur_id) if cur_id else None
    filas   = []
    if cur_id:
        q = _inscritos_curso(cur_id)
        if orden == 'ci':    q = q.order_by(Alumno.ci)
        elif orden == 'sexo':q = q.order_by(Alumno.masculino.desc(), Alumno.paterno)
        else:                q = q.order_by(Alumno.paterno, Alumno.nombre)
        for i, ins in enumerate(q.all(), 1):
            a = ins.alumno
            filas.append({'nro':i,'nombre':a.nombre_completo,'ci':a.ci or '—',
                'nac':a.nacimiento.strftime('%d/%m/%Y') if a.nacimiento else '—',
                'sexo':'M' if a.masculino else 'F','email':a.email or '—',
                'estado':'Inscrito' if ins.inscrito else 'Reserva',
                'desc':f'{ins.descuento}%' if ins.descuento else '—'})
    if formato=='excel': return _r1_xl(filas,curso)
    if formato=='pdf':   return _r1_pdf(filas,curso)
    return render_template('reportes/r1.html', filas=filas, cur_id=cur_id,
                           curso=curso, orden=orden, cursos=_cursos())

def _r1_xl(filas, curso):
    from openpyxl import Workbook
    wb=Workbook(); ws=wb.active; ws.title='Alumnos por Curso'; s=_xl_styles()
    _xl_hdr(ws,['#','Apellidos y Nombre','CI','Nacimiento','Sexo','Email','Estado','Descuento'],s)
    for f in filas: ws.append([f['nro'],f['nombre'],f['ci'],f['nac'],f['sexo'],f['email'],f['estado'],f['desc']])
    _xl_rows(ws,s); _xl_width(ws)
    return _xl_send(wb, f'r1_alumnos_{(_nc(curso) or "todos").replace(" ","_")}.xlsx')

def _r1_pdf(filas, curso):
    from reportlab.lib.units import cm
    doc,buf=_pdf_setup()
    story=_pdf_title('Alumnos por Curso', _nc(curso))
    enc=[['#','Nombre','CI','Nac.','Sexo','Email','Estado','Desc.']]
    data=enc+[[f['nro'],f['nombre'],f['ci'],f['nac'],f['sexo'],f['email'],f['estado'],f['desc']] for f in filas]
    story.append(_pdf_tbl(data,[.5*cm,4.5*cm,1.5*cm,2*cm,1*cm,3.5*cm,1.5*cm,1*cm]))
    return _pdf_send(doc,buf,story,f'r1_alumnos_{(_nc(curso) or "todos").replace(" ","_")}.pdf')


# ═══════════════════════════════════════════════════════════════════════════════
# R2 – Alumnos con Calificaciones
# ═══════════════════════════════════════════════════════════════════════════════

@reportes_bp.route('/r2')
@login_required
def r2():
    cur_id  = request.args.get('cur_id', type=int)
    mat_id  = request.args.get('mat_id', type=int)
    orden   = request.args.get('orden', 'nombre')
    formato = request.args.get('formato', 'html')
    curso   = Curso.query.get(cur_id) if cur_id else None
    materias= []
    filas   = []
    if cur_id:
        mat_ids  = [a.mat_id for a in Asignado.query.filter_by(cur_id=cur_id).all()]
        materias = Materia.query.filter(Materia.id.in_(mat_ids)).order_by(Materia.materia).all()
        q = (db.session.query(Nota,Alumno,Materia)
             .join(Inscrito,Nota.ins_id==Inscrito.id)
             .join(Alumno,Inscrito.alu_id==Alumno.id)
             .join(Materia,Nota.mat_id==Materia.id)
             .filter(Inscrito.cur_id==cur_id))
        if mat_id: q=q.filter(Nota.mat_id==mat_id)
        if orden=='nota':     q=q.order_by(Nota.nota_final.desc(),Alumno.paterno)
        elif orden=='estado': q=q.order_by(Nota.aprobado.desc(),Nota.nota_final.desc())
        else:                 q=q.order_by(Alumno.paterno,Alumno.nombre,Materia.materia)
        for i,(n,a,m) in enumerate(q.all(),1):
            filas.append({'nro':i,'alumno':a.nombre_completo,'materia':m.materia,
                'n1':n.nota1,'n2':n.nota2,'n3':n.nota3,
                'nf':float(n.nota_final),
                'apr':'Aprobado' if n.aprobado else 'Reprobado','obs':n.obs or ''})
    if formato=='excel': return _r2_xl(filas,curso)
    if formato=='pdf':   return _r2_pdf(filas,curso)
    return render_template('reportes/r2.html', filas=filas, cur_id=cur_id,
                           mat_id=mat_id, curso=curso, orden=orden,
                           materias=materias, cursos=_cursos())

def _r2_xl(filas, curso):
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill
    wb=Workbook(); ws=wb.active; ws.title='Calificaciones'; s=_xl_styles()
    _xl_hdr(ws,['#','Alumno','Materia','P1','P2','P3','Promedio','Estado','Observaciones'],s)
    gf=PatternFill('solid',fgColor='FFD4EDDA'); rf=PatternFill('solid',fgColor='FFF8D7DA')
    for f in filas:
        ws.append([f['nro'],f['alumno'],f['materia'],f['n1'],f['n2'],f['n3'],f['nf'],f['apr'],f['obs']])
        fill=gf if f['apr']=='Aprobado' else rf
        for c in ws[ws.max_row]:
            c.fill=fill; c.font=s['bfont']; c.border=s['border']; c.alignment=s['left']
    _xl_width(ws)
    return _xl_send(wb,f'r2_calificaciones_{(_nc(curso) or "todos").replace(" ","_")}.xlsx')

def _r2_pdf(filas, curso):
    from reportlab.lib.units import cm
    doc,buf=_pdf_setup(landscape=True)
    story=_pdf_title('Calificaciones por Curso',_nc(curso))
    enc=[['#','Alumno','Materia','P1','P2','P3','Prom.','Estado','Obs.']]
    data=enc+[[f['nro'],f['alumno'],f['materia'],f['n1'],f['n2'],f['n3'],f['nf'],f['apr'],f['obs']] for f in filas]
    story.append(_pdf_tbl(data,[.5*cm,4.5*cm,3.5*cm,1*cm,1*cm,1*cm,1.2*cm,2*cm,3*cm]))
    return _pdf_send(doc,buf,story,f'r2_calificaciones_{(_nc(curso) or "todos").replace(" ","_")}.pdf')


# ═══════════════════════════════════════════════════════════════════════════════
# R3 – Estado de Pagos por Curso
# ═══════════════════════════════════════════════════════════════════════════════

@reportes_bp.route('/r3')
@login_required
def r3():
    cur_id  = request.args.get('cur_id', type=int)
    orden   = request.args.get('orden', 'nombre')
    formato = request.args.get('formato', 'html')
    curso   = Curso.query.get(cur_id) if cur_id else None
    filas   = []
    if cur_id:
        rows=[]
        for ins in _inscritos_curso(cur_id).order_by(Alumno.paterno,Alumno.nombre).all():
            ps=ins.pagos.all()
            tc=sum(p.cuota for p in ps); pg=sum(p.cuota for p in ps if p.pagado)
            pend=tc-pg; cp=len(ps); cpp=sum(1 for p in ps if p.pagado)
            rows.append({'nombre':ins.alumno.nombre_completo,'ci':ins.alumno.ci or '—',
                'ct':cp,'cpg':cpp,'total':tc,'pagado':pg,'pendiente':pend,
                'estado':'Al día' if pend==0 and cp>0 else 'Con deuda' if pend>0 else 'Sin plan'})
        if orden=='deuda':  rows.sort(key=lambda x:x['pendiente'],reverse=True)
        elif orden=='estado':rows.sort(key=lambda x:x['estado'])
        for i,r in enumerate(rows,1): r['nro']=i; filas.append(r)
    if formato=='excel': return _r3_xl(filas,curso)
    if formato=='pdf':   return _r3_pdf(filas,curso)
    return render_template('reportes/r3.html', filas=filas, cur_id=cur_id,
                           curso=curso, orden=orden, cursos=_cursos())

def _r3_xl(filas,curso):
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill
    wb=Workbook(); ws=wb.active; ws.title='Estado de Pagos'; s=_xl_styles()
    _xl_hdr(ws,['#','Alumno','CI','Cuotas Tot.','Cuotas Pag.','Total Bs.','Pagado Bs.','Pendiente Bs.','Estado'],s)
    df=PatternFill('solid',fgColor='FFF8D7DA'); of=PatternFill('solid',fgColor='FFD4EDDA')
    for f in filas:
        ws.append([f['nro'],f['nombre'],f['ci'],f['ct'],f['cpg'],
                   round(f['total'],2),round(f['pagado'],2),round(f['pendiente'],2),f['estado']])
        fill=of if f['estado']=='Al día' else df
        for c in ws[ws.max_row]:
            c.fill=fill; c.font=s['bfont']; c.border=s['border']; c.alignment=s['left']
    _xl_width(ws)
    return _xl_send(wb,f'r3_pagos_{(_nc(curso) or "todos").replace(" ","_")}.xlsx')

def _r3_pdf(filas,curso):
    from reportlab.lib.units import cm
    doc,buf=_pdf_setup(landscape=True)
    story=_pdf_title('Estado de Pagos por Curso',_nc(curso))
    enc=[['#','Alumno','CI','C.Tot','C.Pag','Total Bs.','Pagado Bs.','Pendiente Bs.','Estado']]
    data=enc+[[f['nro'],f['nombre'],f['ci'],f['ct'],f['cpg'],
               f"{f['total']:.2f}",f"{f['pagado']:.2f}",f"{f['pendiente']:.2f}",f['estado']] for f in filas]
    story.append(_pdf_tbl(data,[.5*cm,4.5*cm,1.5*cm,1.2*cm,1.2*cm,2*cm,2*cm,2.5*cm,2*cm]))
    return _pdf_send(doc,buf,story,f'r3_pagos_{(_nc(curso) or "todos").replace(" ","_")}.pdf')


# ═══════════════════════════════════════════════════════════════════════════════
# R4 – Estado de Cuenta de un Alumno
# ═══════════════════════════════════════════════════════════════════════════════

@reportes_bp.route('/r4')
@login_required
def r4():
    alu_id  = request.args.get('alu_id', type=int)
    formato = request.args.get('formato', 'html')
    alumno  = Alumno.query.get(alu_id) if alu_id else None
    ins     = Inscrito.query.filter_by(alu_id=alu_id).first() if alu_id else None
    pagos=[]; resumen={}
    if ins:
        for p in Pago.query.filter_by(ins_id=ins.id).order_by(Pago.nro_cuota).all():
            pagos.append({'nro':p.nro_cuota,'cuota':p.cuota,'pagado':p.pagado,
                'metodo':p.metodo_pago or '—',
                'fecha':p.fecha_pago.strftime('%d/%m/%Y') if p.fecha_pago else '—',
                'ref':p.referencia_pago or '—','obs':p.obs or ''})
        total=sum(p['cuota'] for p in pagos); pg=sum(p['cuota'] for p in pagos if p['pagado'])
        resumen={'total':total,'pagado':pg,'pendiente':total-pg,
                 'cuotas':len(pagos),'cpag':sum(1 for p in pagos if p['pagado'])}
    if formato=='excel': return _r4_xl(alumno,ins,pagos,resumen)
    if formato=='pdf':   return _r4_pdf(alumno,ins,pagos,resumen)
    return render_template('reportes/r4.html', alumno=alumno, ins=ins,
                           pagos=pagos, resumen=resumen, alu_id=alu_id,
                           alumnos=_alumnos_activos())

def _r4_xl(alumno,ins,pagos,resumen):
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill
    wb=Workbook(); ws=wb.active; ws.title='Estado de Cuenta'; s=_xl_styles()
    ws.append(['Alumno:',alumno.nombre_completo if alumno else ''])
    ws.append(['CI:',str(alumno.ci) if alumno and alumno.ci else ''])
    ws.append(['Curso:',_nc(ins.curso) if ins and ins.curso else ''])
    ws.append(['Descuento:',f'{ins.descuento}%' if ins else '']); ws.append([])
    _xl_hdr(ws,['Cuota #','Monto Bs.','Estado','Método','Fecha Pago','Referencia','Observaciones'],s)
    of=PatternFill('solid',fgColor='FFD4EDDA'); pf=PatternFill('solid',fgColor='FFFFF3CD')
    for p in pagos:
        ws.append([p['nro'],round(p['cuota'],2),'Pagado' if p['pagado'] else 'Pendiente',
                   p['metodo'],p['fecha'],p['ref'],p['obs']])
        fill=of if p['pagado'] else pf
        for c in ws[ws.max_row]:
            c.fill=fill; c.font=s['bfont']; c.border=s['border']
    ws.append([]); ws.append(['','TOTAL:',resumen.get('total',0),'PAGADO:',
                               resumen.get('pagado',0),'PENDIENTE:',resumen.get('pendiente',0)])
    for c in ws[ws.max_row]: c.font=s['boldf']
    _xl_width(ws)
    n=(alumno.nombre_completo.replace(' ','_') if alumno else 'alumno')
    return _xl_send(wb,f'r4_cuenta_{n}.xlsx')

def _r4_pdf(alumno,ins,pagos,resumen):
    from reportlab.lib.units import cm
    from reportlab.platypus import Spacer, Table, TableStyle
    doc,buf=_pdf_setup()
    story=_pdf_title('Estado de Cuenta del Alumno',alumno.nombre_completo if alumno else '')
    info=[['Alumno:',alumno.nombre_completo if alumno else '—'],
          ['CI:',str(alumno.ci) if alumno and alumno.ci else '—'],
          ['Curso:',_nc(ins.curso) if ins and ins.curso else '—'],
          ['Descuento:',f'{ins.descuento}%' if ins and ins.descuento else '0%']]
    it=Table(info,colWidths=[3*cm,12*cm])
    it.setStyle(TableStyle([('FONTNAME',(0,0),(0,-1),'Helvetica-Bold'),
                              ('FONTSIZE',(0,0),(-1,-1),9),
                              ('TOPPADDING',(0,0),(-1,-1),3),
                              ('BOTTOMPADDING',(0,0),(-1,-1),3)]))
    story.append(it); story.append(Spacer(1,.3*cm))
    enc=[['Cuota #','Monto Bs.','Estado','Método','Fecha','Referencia']]
    data=enc+[[p['nro'],f"{p['cuota']:.2f}",'Pagado' if p['pagado'] else 'Pendiente',
               p['metodo'],p['fecha'],p['ref']] for p in pagos]
    story.append(_pdf_tbl(data,[1.5*cm,2.5*cm,2*cm,2.5*cm,2.5*cm,4*cm]))
    story.append(Spacer(1,.3*cm))
    res=[['Total Bs.','Pagado Bs.','Pendiente Bs.','Cuotas','C. Pagadas'],
         [f"{resumen.get('total',0):.2f}",f"{resumen.get('pagado',0):.2f}",
          f"{resumen.get('pendiente',0):.2f}",resumen.get('cuotas',0),resumen.get('cpag',0)]]
    story.append(_pdf_tbl(res,[3*cm,3*cm,3.5*cm,2*cm,3.5*cm],zebra=False))
    n=(alumno.nombre_completo.replace(' ','_') if alumno else 'alumno')
    return _pdf_send(doc,buf,story,f'r4_cuenta_{n}.pdf')


# ═══════════════════════════════════════════════════════════════════════════════
# R5 – Top 3 mejores promedios
# ═══════════════════════════════════════════════════════════════════════════════

@reportes_bp.route('/r5')
@login_required
def r5():
    cur_id  = request.args.get('cur_id', type=int)
    formato = request.args.get('formato', 'html')
    cursos_sel = [Curso.query.get(cur_id)] if cur_id else _cursos()
    filas=[]
    for curso in cursos_sel:
        if not curso: continue
        rows=(db.session.query(
                  Inscrito.id.label('iid'),
                  func.avg(Nota.nota_final).label('prom'),
                  func.sum(db.case((Nota.aprobado==True,1),else_=0)).label('apr'),
                  func.sum(db.case((Nota.aprobado==False,1),else_=0)).label('rep')
              ).join(Nota,Nota.ins_id==Inscrito.id)
              .filter(Inscrito.cur_id==curso.id)
              .group_by(Inscrito.id)
              .order_by(func.avg(Nota.nota_final).desc())
              .limit(3).all())
        for pos,row in enumerate(rows,1):
            ins=Inscrito.query.get(row.iid)
            if ins and ins.alumno:
                filas.append({'curso':_nc(curso),'pos':pos,
                    'alumno':ins.alumno.nombre_completo,
                    'prom':round(float(row.prom),2),
                    'apr':row.apr,'rep':row.rep})
    if formato=='excel': return _r5_xl(filas)
    if formato=='pdf':   return _r5_pdf(filas)
    return render_template('reportes/r5.html', filas=filas,
                           cur_id=cur_id, cursos=_cursos())

def _r5_xl(filas):
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill
    wb=Workbook(); ws=wb.active; ws.title='Top 3'; s=_xl_styles()
    medals={'1':'FFD700','2':'C0C0C0','3':'CD7F32'}
    _xl_hdr(ws,['Curso','Puesto','Alumno','Promedio','Mat. Aprobadas','Mat. Reprobadas'],s)
    for f in filas:
        ws.append([f['curso'],f'#{f["pos"]}',f['alumno'],f['prom'],f['apr'],f['rep']])
        color=medals.get(str(f['pos']),'FFFFFF')
        fill=PatternFill('solid',fgColor=f'FF{color}')
        for c in ws[ws.max_row]:
            c.fill=fill; c.font=s['bfont']; c.border=s['border']; c.alignment=s['left']
    _xl_width(ws)
    return _xl_send(wb,'r5_top3_promedios.xlsx')

def _r5_pdf(filas):
    from reportlab.lib.units import cm
    doc,buf=_pdf_setup()
    story=_pdf_title('Top 3 Mejores Promedios por Curso')
    enc=[['Curso','Puesto','Alumno','Promedio','Aprob.','Reprobadas']]
    data=enc+[[f['curso'],f'#{f["pos"]}',f['alumno'],f['prom'],f['apr'],f['rep']] for f in filas]
    story.append(_pdf_tbl(data,[5*cm,1.2*cm,4.5*cm,1.8*cm,1.5*cm,1.5*cm]))
    return _pdf_send(doc,buf,story,'r5_top3_promedios.pdf')


# ═══════════════════════════════════════════════════════════════════════════════
# R6 – Rendimiento visual (datos para Chart.js)
# ═══════════════════════════════════════════════════════════════════════════════

@reportes_bp.route('/r6')
@login_required
def r6():
    cur_id  = request.args.get('cur_id', type=int)
    formato = request.args.get('formato', 'html')
    curso   = Curso.query.get(cur_id) if cur_id else None
    barras={}; aprob={}; dist={}
    if cur_id:
        rows=(db.session.query(
                  Materia.materia,
                  func.avg(Nota.nota_final).label('prom'),
                  func.sum(db.case((Nota.aprobado==True,1),else_=0)).label('apr'),
                  func.sum(db.case((Nota.aprobado==False,1),else_=0)).label('rep'),
              ).join(Nota,Nota.mat_id==Materia.id)
              .join(Inscrito,Nota.ins_id==Inscrito.id)
              .filter(Inscrito.cur_id==cur_id)
              .group_by(Materia.materia)
              .order_by(Materia.materia).all())
        labels=[r.materia for r in rows]
        barras={'labels':labels,'data':[round(float(r.prom),1) for r in rows]}
        aprob={'labels':labels,
               'apr':[int(r.apr) for r in rows],
               'rep':[int(r.rep) for r in rows]}
        notas=[float(n) for (n,) in
               Nota.query.join(Inscrito,Nota.ins_id==Inscrito.id)
               .filter(Inscrito.cur_id==cur_id)
               .with_entities(Nota.nota_final).all()]
        d=[0,0,0,0]
        for v in notas:
            if v<=50: d[0]+=1
            elif v<=70: d[1]+=1
            elif v<=85: d[2]+=1
            else: d[3]+=1
        dist={'labels':['0–50 Reprobado','51–70 Suficiente','71–85 Bueno','86–100 Excelente'],'data':d}
    if formato=='excel': return _r6_xl(barras,aprob,dist,curso)
    return render_template('reportes/r6.html', cur_id=cur_id, curso=curso,
                           barras=barras, aprob=aprob, dist=dist, cursos=_cursos())

def _r6_xl(barras,aprob,dist,curso):
    from openpyxl import Workbook
    wb=Workbook(); s=_xl_styles()
    ws1=wb.active; ws1.title='Prom. por Materia'
    _xl_hdr(ws1,['Materia','Promedio','Aprobados','Reprobados'],s)
    if barras.get('labels'):
        for i,m in enumerate(barras['labels']):
            ws1.append([m,barras['data'][i],aprob['apr'][i],aprob['rep'][i]])
    _xl_rows(ws1,s); _xl_width(ws1)
    ws2=wb.create_sheet('Distribución')
    _xl_hdr(ws2,['Rango','Cantidad'],s)
    if dist.get('labels'):
        for l,c in zip(dist['labels'],dist['data']): ws2.append([l,c])
    _xl_rows(ws2,s); _xl_width(ws2)
    n=(_nc(curso) or 'todos').replace(' ','_')
    return _xl_send(wb,f'r6_rendimiento_{n}.xlsx')


# ═══════════════════════════════════════════════════════════════════════════════
# R7 – Promedios por Materia
# ═══════════════════════════════════════════════════════════════════════════════

@reportes_bp.route('/r7')
@login_required
def r7():
    cur_id  = request.args.get('cur_id', type=int)
    orden   = request.args.get('orden', 'materia')
    formato = request.args.get('formato', 'html')
    curso   = Curso.query.get(cur_id) if cur_id else None
    filas=[]
    q=(db.session.query(Materia.materia,Curso.paralelo,Curso.gestion,
                         func.avg(Nota.nota_final).label('prom'),
                         func.count(Nota.id).label('tot'),
                         func.sum(db.case((Nota.aprobado==True,1),else_=0)).label('apr'))
       .join(Nota,Nota.mat_id==Materia.id)
       .join(Inscrito,Nota.ins_id==Inscrito.id)
       .join(Curso,Inscrito.cur_id==Curso.id))
    if cur_id: q=q.filter(Curso.id==cur_id)
    q=q.group_by(Materia.materia,Curso.paralelo,Curso.gestion)
    q=q.order_by(func.avg(Nota.nota_final).desc() if orden=='promedio' else Materia.materia)
    for i,row in enumerate(q.all(),1):
        pct=round(float(row.apr)/row.tot*100,1) if row.tot else 0
        filas.append({'nro':i,'materia':row.materia,'curso':f'{row.paralelo} ({row.gestion})',
                      'prom':round(float(row.prom),2),'tot':row.tot,'apr':row.apr,'pct':pct})
    if formato=='excel': return _r7_xl(filas,curso)
    if formato=='pdf':   return _r7_pdf(filas,curso)
    return render_template('reportes/r7.html', filas=filas, cur_id=cur_id,
                           curso=curso, orden=orden, cursos=_cursos())

def _r7_xl(filas,curso):
    from openpyxl import Workbook
    wb=Workbook(); ws=wb.active; ws.title='Prom. por Materia'; s=_xl_styles()
    _xl_hdr(ws,['#','Materia','Curso','Promedio','Total','Aprobados','% Aprobación'],s)
    for f in filas: ws.append([f['nro'],f['materia'],f['curso'],f['prom'],f['tot'],f['apr'],f['pct']])
    _xl_rows(ws,s); _xl_width(ws)
    return _xl_send(wb,f'r7_prom_materia_{(_nc(curso) or "todos").replace(" ","_")}.xlsx')

def _r7_pdf(filas,curso):
    from reportlab.lib.units import cm
    doc,buf=_pdf_setup()
    story=_pdf_title('Promedios por Materia',_nc(curso) or 'Todos los cursos')
    enc=[['#','Materia','Curso','Prom.','Total','Aprob.','% Aprob.']]
    data=enc+[[f['nro'],f['materia'],f['curso'],f['prom'],f['tot'],f['apr'],f['pct']] for f in filas]
    story.append(_pdf_tbl(data,[.6*cm,4.5*cm,3*cm,1.5*cm,1.3*cm,1.3*cm,1.8*cm]))
    return _pdf_send(doc,buf,story,f'r7_prom_materia_{(_nc(curso) or "todos").replace(" ","_")}.pdf')


# ═══════════════════════════════════════════════════════════════════════════════
# R8 – Promedios por Profesor
# ═══════════════════════════════════════════════════════════════════════════════

@reportes_bp.route('/r8')
@login_required
def r8():
    cur_id  = request.args.get('cur_id', type=int)
    orden   = request.args.get('orden', 'profesor')
    formato = request.args.get('formato', 'html')
    curso   = Curso.query.get(cur_id) if cur_id else None
    filas=[]
    q=(db.session.query(
           Profesor.nombre,Profesor.paterno,Profesor.materno,
           Materia.materia,Curso.paralelo,Curso.gestion,
           func.avg(Nota.nota_final).label('prom'),
           func.count(Nota.id).label('tot'),
           func.sum(db.case((Nota.aprobado==True,1),else_=0)).label('apr'))
       .join(Asignado,Asignado.pro_id==Profesor.id)
       .join(Materia,Asignado.mat_id==Materia.id)
       .join(Curso,Asignado.cur_id==Curso.id)
       .join(Inscrito,Inscrito.cur_id==Curso.id)
       .join(Nota,db.and_(Nota.ins_id==Inscrito.id,Nota.mat_id==Materia.id)))
    if cur_id: q=q.filter(Curso.id==cur_id)
    q=q.group_by(Profesor.nombre,Profesor.paterno,Profesor.materno,
                  Materia.materia,Curso.paralelo,Curso.gestion)
    q=q.order_by(func.avg(Nota.nota_final).desc() if orden=='promedio'
                  else (Profesor.paterno,Profesor.nombre))
    for i,row in enumerate(q.all(),1):
        np=f'{row.nombre} {row.paterno or ""} {row.materno or ""}'.strip()
        pct=round(float(row.apr)/row.tot*100,1) if row.tot else 0
        filas.append({'nro':i,'prof':np,'materia':row.materia,
                      'curso':f'{row.paralelo} ({row.gestion})',
                      'prom':round(float(row.prom),2),'tot':row.tot,'apr':row.apr,'pct':pct})
    if formato=='excel': return _r8_xl(filas,curso)
    if formato=='pdf':   return _r8_pdf(filas,curso)
    return render_template('reportes/r8.html', filas=filas, cur_id=cur_id,
                           curso=curso, orden=orden, cursos=_cursos())

def _r8_xl(filas,curso):
    from openpyxl import Workbook
    wb=Workbook(); ws=wb.active; ws.title='Prom. por Profesor'; s=_xl_styles()
    _xl_hdr(ws,['#','Profesor','Materia','Curso','Promedio','Total','Aprobados','% Aprobación'],s)
    for f in filas: ws.append([f['nro'],f['prof'],f['materia'],f['curso'],f['prom'],f['tot'],f['apr'],f['pct']])
    _xl_rows(ws,s); _xl_width(ws)
    return _xl_send(wb,f'r8_prom_profesor_{(_nc(curso) or "todos").replace(" ","_")}.xlsx')

def _r8_pdf(filas,curso):
    from reportlab.lib.units import cm
    doc,buf=_pdf_setup(landscape=True)
    story=_pdf_title('Promedios por Profesor',_nc(curso) or 'Todos los cursos')
    enc=[['#','Profesor','Materia','Curso','Prom.','Total','Aprob.','% Aprob.']]
    data=enc+[[f['nro'],f['prof'],f['materia'],f['curso'],f['prom'],f['tot'],f['apr'],f['pct']] for f in filas]
    story.append(_pdf_tbl(data,[.5*cm,4*cm,3.5*cm,2.5*cm,1.5*cm,1.2*cm,1.2*cm,1.8*cm]))
    return _pdf_send(doc,buf,story,f'r8_prom_profesor_{(_nc(curso) or "todos").replace(" ","_")}.pdf')


# ═══════════════════════════════════════════════════════════════════════════════
# R9 – Alumnos con Abandono / Deuda  (sugerido)
# ═══════════════════════════════════════════════════════════════════════════════

@reportes_bp.route('/r9')
@login_required
def r9():
    tipo    = request.args.get('tipo','deuda')
    cur_id  = request.args.get('cur_id', type=int)
    formato = request.args.get('formato','html')
    curso   = Curso.query.get(cur_id) if cur_id else None
    filas=[]
    q=Inscrito.query.join(Alumno,Inscrito.alu_id==Alumno.id)
    if cur_id: q=q.filter(Inscrito.cur_id==cur_id)
    for ins in q.order_by(Alumno.paterno,Alumno.nombre).all():
        ps=ins.pagos.all(); pend=sum(p.cuota for p in ps if not p.pagado)
        ed=pend>0; ea=ins.abandono
        if tipo=='deuda'    and not ed: continue
        if tipo=='abandono' and not ea: continue
        if tipo=='ambos'    and not (ed or ea): continue
        filas.append({'alumno':ins.alumno.nombre_completo,'ci':ins.alumno.ci or '—',
            'curso':_nc(ins.curso) if ins.curso else '—',
            'deuda':round(pend,2),'abandono':'Sí' if ea else 'No','obs':ins.obs or ''})
    if formato=='excel': return _r9_xl(filas,tipo)
    if formato=='pdf':   return _r9_pdf(filas,tipo)
    return render_template('reportes/r9.html', filas=filas, cur_id=cur_id,
                           tipo=tipo, curso=curso, cursos=_cursos())

def _r9_xl(filas,tipo):
    from openpyxl import Workbook
    wb=Workbook(); ws=wb.active; ws.title='Abandono/Deudores'; s=_xl_styles()
    _xl_hdr(ws,['Alumno','CI','Curso','Deuda Bs.','Abandono','Obs.'],s)
    for f in filas: ws.append([f['alumno'],f['ci'],f['curso'],f['deuda'],f['abandono'],f['obs']])
    _xl_rows(ws,s); _xl_width(ws)
    return _xl_send(wb,f'r9_{tipo}.xlsx')

def _r9_pdf(filas,tipo):
    from reportlab.lib.units import cm
    doc,buf=_pdf_setup()
    story=_pdf_title('Alumnos con Deuda / Abandono')
    enc=[['Alumno','CI','Curso','Deuda Bs.','Abandono','Obs.']]
    data=enc+[[f['alumno'],f['ci'],f['curso'],f'{f["deuda"]:.2f}',f['abandono'],f['obs']] for f in filas]
    story.append(_pdf_tbl(data,[4.5*cm,1.5*cm,3.5*cm,2*cm,1.5*cm,3*cm]))
    return _pdf_send(doc,buf,story,f'r9_{tipo}.pdf')


# ═══════════════════════════════════════════════════════════════════════════════
# R10 – Resumen Financiero por Curso  (sugerido)
# ═══════════════════════════════════════════════════════════════════════════════

@reportes_bp.route('/r10')
@login_required
def r10():
    formato = request.args.get('formato','html')
    filas=[]
    for c in _cursos():
        ins_list=(Inscrito.query
                  .filter(Inscrito.cur_id==c.id,
                          db.or_(Inscrito.inscrito==True,Inscrito.reserva==True)).all())
        esp=sum(sum(p.cuota for p in i.pagos.all()) for i in ins_list)
        rec=sum(sum(p.cuota for p in i.pagos.all() if p.pagado) for i in ins_list)
        pend=esp-rec; pct=round(rec/esp*100,1) if esp else 0
        filas.append({'curso':_nc(c),'alumnos':len(ins_list),
                      'esp':round(esp,2),'rec':round(rec,2),
                      'pend':round(pend,2),'pct':pct})
    if formato=='excel': return _r10_xl(filas)
    if formato=='pdf':   return _r10_pdf(filas)
    return render_template('reportes/r10.html', filas=filas)

def _r10_xl(filas):
    from openpyxl import Workbook
    wb=Workbook(); ws=wb.active; ws.title='Resumen Financiero'; s=_xl_styles()
    _xl_hdr(ws,['Curso','Alumnos','Total Esp. Bs.','Recaudado Bs.','Pendiente Bs.','% Cobrado'],s)
    for f in filas: ws.append([f['curso'],f['alumnos'],f['esp'],f['rec'],f['pend'],f['pct']])
    _xl_rows(ws,s); _xl_width(ws)
    return _xl_send(wb,'r10_resumen_financiero.xlsx')

def _r10_pdf(filas):
    from reportlab.lib.units import cm
    doc,buf=_pdf_setup()
    story=_pdf_title('Resumen Financiero por Curso')
    enc=[['Curso','Alumnos','Esperado Bs.','Recaudado Bs.','Pendiente Bs.','% Cobrado']]
    data=enc+[[f['curso'],f['alumnos'],f'{f["esp"]:.2f}',f'{f["rec"]:.2f}',f'{f["pend"]:.2f}',f'{f["pct"]}%'] for f in filas]
    story.append(_pdf_tbl(data,[4.5*cm,1.8*cm,3*cm,3*cm,3*cm,2*cm]))
    return _pdf_send(doc,buf,story,'r10_resumen_financiero.pdf')


# ═══════════════════════════════════════════════════════════════════════════════
# R11 – Tasa de Aprobación por Materia  (sugerido)
# ═══════════════════════════════════════════════════════════════════════════════

@reportes_bp.route('/r11')
@login_required
def r11():
    cur_id  = request.args.get('cur_id', type=int)
    orden   = request.args.get('orden','materia')
    formato = request.args.get('formato','html')
    curso   = Curso.query.get(cur_id) if cur_id else None
    filas=[]
    q=(db.session.query(Materia.materia,
                         func.count(Nota.id).label('tot'),
                         func.sum(db.case((Nota.aprobado==True,1),else_=0)).label('apr'),
                         func.avg(Nota.nota_final).label('prom'),
                         func.min(Nota.nota_final).label('mn'),
                         func.max(Nota.nota_final).label('mx'))
       .join(Nota,Nota.mat_id==Materia.id)
       .join(Inscrito,Nota.ins_id==Inscrito.id))
    if cur_id: q=q.filter(Inscrito.cur_id==cur_id)
    q=q.group_by(Materia.materia)
    q=q.order_by((func.sum(db.case((Nota.aprobado==True,1),else_=0))/func.count(Nota.id)).desc()
                  if orden=='tasa' else Materia.materia)
    for i,row in enumerate(q.all(),1):
        rep=row.tot-row.apr; pct=round(float(row.apr)/row.tot*100,1) if row.tot else 0
        filas.append({'nro':i,'materia':row.materia,'tot':row.tot,'apr':row.apr,
                      'rep':rep,'pct':pct,'prom':round(float(row.prom),2),
                      'mn':float(row.mn),'mx':float(row.mx)})
    if formato=='excel': return _r11_xl(filas,curso)
    if formato=='pdf':   return _r11_pdf(filas,curso)
    return render_template('reportes/r11.html', filas=filas, cur_id=cur_id,
                           curso=curso, orden=orden, cursos=_cursos())

def _r11_xl(filas,curso):
    from openpyxl import Workbook
    wb=Workbook(); ws=wb.active; ws.title='Tasa Aprobación'; s=_xl_styles()
    _xl_hdr(ws,['#','Materia','Total','Aprob.','Reprobados','% Aprob.','Prom.','Mín.','Máx.'],s)
    for f in filas: ws.append([f['nro'],f['materia'],f['tot'],f['apr'],f['rep'],f['pct'],f['prom'],f['mn'],f['mx']])
    _xl_rows(ws,s); _xl_width(ws)
    return _xl_send(wb,f'r11_tasa_aprobacion_{(_nc(curso) or "todos").replace(" ","_")}.xlsx')

def _r11_pdf(filas,curso):
    from reportlab.lib.units import cm
    doc,buf=_pdf_setup(landscape=True)
    story=_pdf_title('Tasa de Aprobación por Materia',_nc(curso) or 'Todos los cursos')
    enc=[['#','Materia','Total','Aprob.','Reprobados','% Aprob.','Prom.','Mín.','Máx.']]
    data=enc+[[f['nro'],f['materia'],f['tot'],f['apr'],f['rep'],f'{f["pct"]}%',f['prom'],f['mn'],f['mx']] for f in filas]
    story.append(_pdf_tbl(data,[.6*cm,4*cm,1.2*cm,1.3*cm,1.5*cm,1.5*cm,2*cm,1.2*cm,1.2*cm]))
    return _pdf_send(doc,buf,story,f'r11_tasa_aprobacion_{(_nc(curso) or "todos").replace(" ","_")}.pdf')
